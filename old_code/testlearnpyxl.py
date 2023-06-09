from openpyxl import Workbook
from openpyxl import load_workbook

'''
This file can be used to learn openpyxl
'''


wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("mySheet1") # insert at the end (default)
# or
ws2 = wb.create_sheet("mysheet2") # insert at first position(,0)
# or
ws3 = wb.create_sheet("Mysheet3") # insert at the penultimate position(,-1)

date = "4-25_5-1"
ws.title = "mainSheet"
ws1.title = f"{date}_info"
# ws2.title = "mysheet"

ws1 = wb[f"{date}_info"] #gets key of workbook

# sheets can be iterated through like this
# for sheet in wb:
#     print(sheet.title)

# how to access specific cells, returns cell- if doesn't exist
# will make one


nameCell = ws['A4']

# or

# access to cells with row/col notation
name = ws.cell(row=4,column=2,value=100)

# when a worksheet is created in memory, it contains no cells. They are created when first accessed.

# cells can be accessed in a range
cell_range = ws['A1':'C2']

# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
#     for cell in row:
#         print(cell)

# for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
#     for cell in col:
#         print(cell)


# how to open an existing workbook


myWb = load_workbook(filename='tradingpost.xlsx')
myWs = myWb.active

cell_tickers = myWs['A']

# for i in cell_tickers:
#     print(i)

# for cell in myWs:
#     print(cell)

# for row in myWs.values:
#     for value in row:
#         print(value)

# indexing for all tickers on scorecard to get values
for col in myWs.iter_rows(min_row=5, max_col=1, max_row=30, values_only=True):
    print(col)