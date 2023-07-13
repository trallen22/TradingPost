'''
This file holds all the configuraion information
'''
from datetime import date, timedelta
from polygon import RESTClient 
import os
import sys 
import openpyxl


# Email variables 
EMAILADDRESS = 'etfsender@gmail.com'
EMAILPASSWORD = 'egztwpmmkbicpjfd' # 'P@55w0rd123' 
EMAILLIST = [ 'trallen@davidson.edu' ]

SRCPATH = os.path.dirname(__file__) + '/'

# convert todays date to mm/dd form 
today = date.today() - timedelta(1)

listDate = str(today).split('-')
TODAYDATE = f'{listDate[1]}/{listDate[2]}' 

STRTODAY = today.strftime('%Y-%m-%d') # used with polygon data

PRINTDF = 0 # prints dataframes to terminal
PBAR = 1 # print progress bar for polygon calls in generate_csv.py 
DEBUG = 1 # print debug messages # TODO20: add debug messages 
DEBUGDATA = 0 # print debug messages from get_data.py
CSV = 1 # outputs an excel file to CSVFILE 
FILLPLATFORM = 1
SENDEMAIL = 1

for arg in sys.argv:
        if arg == '-h':
                print('Need to work on help menu')
                exit(0)
        if arg == '-f':
                PRINTDF = 1
        if arg == '-p':
                PBAR = 1 
        if arg == '-d':
                DEBUG = 1 
        if arg == '-c':
                CSV = 1
        if arg == '-m':
                FILLPLATFORM = 1 
        if arg == '-e':
                SENDEMAIL = 1

# polygon login 
'''Insert your key. Play around with the free tier key first.'''
key = "nGJdIcDOy3hzWwn6X6gritFJkgDWTpRJ"
try:
        winebagle = 1
        while (winebagle):
                print(winebagle)
                CLIENT = RESTClient(key)
                print(f'no win')
                winebagle = 0
except Exception:
        print('failed to connect to Polygon rest client')

PARAMSET = [[ 'minute', 1 ], # one minute time interval 
                [ 'minute', 5 ], # 5 minute time interval 
                [ 'day', 1 ]] #one day time interval 

# Platform files 
TEMPLATEPLATFORM = SRCPATH + 'TA.WORK.xlsx'
OUTPUTPLATFORM = f'/Users/tristanallen/Desktop/TradingPost/testTP/{listDate[1]}-{listDate[2]}_testPlatform.xlsx'
RAWPLATFORM = SRCPATH + 'rawPlatform.xlsx'

# Trading Post files
TEMPEXCEL = SRCPATH + 'stocktradingpost.xlsx'
OUTPUTEXCEL = f'/Users/tristanallen/Desktop/TradingPost/testTP/{listDate[1]}-{listDate[2]}_testTradingPost.xlsx'

CSVFILE = SRCPATH + 'testCsv.csv' 

# TICKERS = [ 'JNK', 'GDX', 'VCR', 'VDC', 'VIG', 'VDE', 'VFH', 
#         'VWO', 'VHT', 'VIS', 'VGT', 'VAW', 'VNQ', 'VOO', 
#         'VOX', 'BND', 'BNDX', 'VXUS', 'VTI', 'VPU', 'XTN' ]

TICKERS = [ 'JNK' ] # used for testing 

INDICATORS = [ 'one_min_50', 'one_min_200', 'five_min_50', 'five_min_200', 'one_day_50', 'one_day_200', 'close_price' ]
MINDICATORS = [ 'five_min_50', 'five_min_200', 'one_min_50', 'one_min_200' ]
DAYDICATORS = [ 'one_day_50', 'one_day_200' ]

INPUTS = { 'G':'close_price', 'H':'one_day_50', 'I':'one_day_200', 'J':'five_min_50', 
        'K':'five_min_200', 'L':'one_min_50', 'M':'one_min_200' }

PLATFORMCOLS = { 'date':'E', 'close_price':'G', 'one_day_50':'H', 'one_day_200':'I', 'five_min_50':'J', 
        'five_min_200':'K', 'one_min_50':'L', 'one_min_200':'M' }

ETFBASECELL = { 'JNK':'C7', 'GDX':'D7', 'VCR':'E7', 'VDC':'F7', 'VIG':'G7', 
        'VDE':'H7', 'VFH':'I7', 'VWO':'C17', 'VHT':'D17', 'VIS':'E17', 'VGT':'F17', 
        'VAW':'G17', 'VNQ':'H17', 'VOO':'I17', 'VOX':'C27', 'BND':'D27', 
        'BNDX':'E27', 'VXUS':'F27', 'VTI':'G27', 'VPU':'H27', 'XTN':'I27' }

PLATDATECELL = 'B3'

try: 
        # loading excel as workbook object
        workbook = openpyxl.load_workbook(TEMPEXCEL)
        excelSheet = workbook.active
except Exception as e:
        print(f'ERROR: {e}')
        exit(4)

COLORROW = 5 # row on trading post excel with template color
plainRGB = 'FFFFFFFF'
PLAINCOLOR = openpyxl.styles.PatternFill(start_color=plainRGB, end_color=plainRGB, fill_type='solid')

# Cell color templates 
try:
        buyRGB = excelSheet[f'A{COLORROW}'].fill.fgColor
        BUYCOLOR = openpyxl.styles.PatternFill(start_color=buyRGB, end_color=buyRGB, fill_type='solid')
except Exception:
        BUYCOLOR = PLAINCOLOR
        print(f'NOTICE: Buy Color set as plain')
try:
        hoBuyRGB = excelSheet[F'C{COLORROW}'].fill.fgColor
        HOBUYCOLOR = openpyxl.styles.PatternFill(start_color=hoBuyRGB, end_color=hoBuyRGB, fill_type='solid')
except Exception:
        HOBUYCOLOR = PLAINCOLOR
        print(f'NOTICE: Hold to Buy Color set as plain')
try:
        sellRGB = excelSheet[f'E{COLORROW}'].fill.fgColor
        SELLCOLOR = openpyxl.styles.PatternFill(start_color=sellRGB, end_color=sellRGB, fill_type='solid')
except Exception:
        SELLCOLOR = PLAINCOLOR
        print(f'NOTICE: Sell Color set as plain')
try:
        hoSellRGB = excelSheet[f'F{COLORROW}'].fill.fgColor
        HOSELLCOLOR = openpyxl.styles.PatternFill(start_color=hoSellRGB, end_color=hoSellRGB, fill_type='solid')
except Exception:
        HOSELLCOLOR = PLAINCOLOR
        print(f'NOTICE: Hold to Sell Color set as plain')

