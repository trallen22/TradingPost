'''
This script copies the template excel file from configurationFile.py and 
creates the new excel for current day based on generated csv

log numbers 300-399
'''

import shutil
import openpyxl 
import configuration_file as config
import plat_formulas as formula


# TODO19: Implement set_ranges 
def set_ranges(rangeType, etf):
    if rangeType == 'buy':
        rangeMin = formula.buy_min(etf)
        rangeMax = formula.buy_max(etf)
    elif rangeType == 'sell':
        rangeMin = formula.sell_min(etf)
        rangeMax = formula.sell_max(etf)
    else:
        return '', '', 1
    if rangeMin > rangeMax:
        return '', '', 1
    return rangeMin, rangeMax, 0

def determine_buy_sell(etf):
    etfVals = etf.indicatorDict
    signal = None
    color = None
    minTradeRange = None
    maxTradeRange = None 

    if (etfVals['close_price'] > etfVals['one_day_50']):
        # Looking for sell signals 
        if (etfVals['close_price'] > etfVals['one_day_200']):
            minPrice = 1000000000
            for col in config.MINDICATORS:
                minPrice = min(minPrice, etfVals[col])
            if (etfVals['close_price'] < minPrice):
                signal = '!SELL!' 
                color = config.SELLCOLOR
            else:
                signal = 'HOLD' 
                color = config.HOSELLCOLOR
            minTradeRange, maxTradeRange, clearTP = set_ranges('sell', etf)
            if (not clearTP):
                return signal, color, minTradeRange, maxTradeRange
    else:
        # Looking for buy signals 
        if (etfVals['close_price'] < etfVals['one_day_50']):
            maxPrice = -1
            for col in config.MINDICATORS:
                maxPrice = max(maxPrice, etfVals[col])
            if (etfVals['close_price'] > maxPrice):
                signal = '!BUY!' 
                color = config.BUYCOLOR 
            else:
                signal = 'HOLD' 
                color = config.HOBUYCOLOR 
            minTradeRange, maxTradeRange, clearTP = set_ranges('buy', etf)
            if (not clearTP):
                return signal, color, minTradeRange, maxTradeRange
    signal = 'HOLD' 
    color = config.PLAINCOLOR 
    minTradeRange, maxTradeRange, clearTP = set_ranges('', etf)
    return signal, color, minTradeRange, maxTradeRange

# TODO: move this function to generate_files.py 
def fill_platform(etfDict, outputPlatform):

    tickerRowDict = {} # { 'JNK': 6 }
    
    # makes a copy of the template platform file
    try:
        shutil.copyfile(config.TEMPLATEPLATFORM, outputPlatform)
        config.logmsg('DEBUG', 300, f'copied template platform to {outputPlatform}')
    except Exception as e:
        config.logmsg('ERROR', 301, f'{e}')
        return 1

    # loading excel as workbook object
    try:
        workbook = openpyxl.load_workbook(outputPlatform)
        config.logmsg('DEBUG', 302, f'loaded output platform as excel')
    except Exception as e:
        config.logmsg('ERROR', 303, f'{e}')
        return 1

    activeSheet = workbook.active

    # setting the platform date cell 
    activeSheet[config.PLATDATECELL] = f'{config.TODAYDATE}/{config.listDate[0]}'

    # going through each cell and getting ticker index 
    config.logmsg('DEBUG', 305, 'getting platform rows for each ticker')
    for row in activeSheet.iter_rows(max_row=activeSheet.max_row, max_col=1):
        for cell in row:
            # gets the rows with tickers in the excel 
            if cell.value in config.TICKERS:
                tickerRowDict[cell.value] = cell.coordinate[1:]

    for ticker in config.TICKERS:
        curRow = tickerRowDict[ticker]
        config.logmsg('DEBUG', 330 + config.TICKERS.index(ticker), f'filling platform, row: {curRow} for ticker: {ticker}')
        for col in config.PLATFORMCOLS:
            if col == 'date':
                activeSheet[f'{config.PLATFORMCOLS[col]}{curRow}'] = config.TODAYDATE
            else:
                activeSheet[f'{config.PLATFORMCOLS[col]}{curRow}'] = etfDict[ticker].indicatorDict[col]

    try:
        workbook.save(outputPlatform)
        workbook.close()
        config.logmsg('DEBUG', 304, 'saving output platform')
    except Exception as e:
        config.logmsg('ERROR', 305, f'{e}')
        return 1

    return 0
