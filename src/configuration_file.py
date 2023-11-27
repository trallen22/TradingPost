'''
This file holds all the configuraion information

log numbers 001-099
'''

from datetime import date, timedelta 
import time 
from polygon import RESTClient 
import os
import sys 
import openpyxl
import argparse 

# logmsg: logs a given message to the log file at LOGFILE
# parameters: 
#       level - string, log level for message urgency -> 'ERROR', 'INFO', 'NOTICE', 'DEBUG' 
#       logNum - int/string, log ID number for current message
#       message - string, message to log 
# returns:  No returns 
def logmsg(level, logNum, message):
    localTime = time.localtime()
    curTime = time.strftime("%H:%M:%S", localTime)
    logMessage = f'{date.today()}::{curTime}::{level}::{logNum}::{message}'
    # prints messages that aren't DEBUG 
    if not (level == 'DEBUG'): 
        print(logMessage)
    try: 
        # writes output to LOGFILE, checks if DEBUG is enabled 
        with open(LOGFILE, mode='a') as logFile:
            if not (level == 'DEBUG' and not DEBUG):
                logFile.write(f'{logMessage}\n')
    except Exception as e:
        print(f'{e}')

# get_args: gets the command line arguments that were given 
# parameters: no parameters 
# returns: returns a dictionary of command line constants and their values 
def get_args():
    parser = argparse.ArgumentParser(description='generates a Trading Post')
    # adding each argument to the parser 
    parser.add_argument('-f', '--PRINTDF', action='store_true', help='prints dataframes to terminal') 
    parser.add_argument('-p', '--PBAR', action='store_true', default=True, help='print progress bar for polygon calls in main.py') 
    parser.add_argument('-d', '--DEBUG', action='store_true', help='log debug messages to LOGFILE')
    parser.add_argument('-c', '--CSV', action='store_true', help='outputs an excel file to CSVFILE') 
    parser.add_argument('-m', '--FILLPLATFORM', action='store_true', help='outputs a platform to OUTPUTPLATFORM')
    parser.add_argument('-x', '--TP', action='store_true', help='generates an excel trading post')
    parser.add_argument('-e', '--SENDEMAIL', action='store_true', help='sends an email to EMAILLIST')
    parser.add_argument('-v', '--GETVALUE', action='store_true', help='gets a specific value from given date. NEED TO IMPLEMENT ') # TODO: implement this
    parser.add_argument('-g', '--GENALL', action='store_true', help='generates csv and platform') 
    parser.add_argument('-t', '--ALTTODAY', metavar='<date>', nargs=1, action='store', default='', help='get Trading Post for specific date. yyyy-mm-dd') 

    args, unknown = parser.parse_known_args()
    return vars(args) 

##############################
# Main Execution begins here 
##############################

argDict = get_args()

PRINTDF = argDict['PRINTDF'] # prints dataframes to terminal
PBAR = argDict['PBAR'] # print progress bar for polygon calls in main.py
DEBUG = argDict['DEBUG'] # log debug messages to LOGFILE 
CSV = argDict['CSV'] # outputs an excel file to CSVFILE 
TP = argDict['TP'] # generates an excel trading post 
FILLPLATFORM = argDict['FILLPLATFORM'] # outputs a platform to OUTPUTPLATFORM 
if argDict['SENDEMAIL']:
    SENDEMAIL = True
    TP = True 
SENDEMAIL = argDict['SENDEMAIL'] # sends an email to EMAILLIST 
GETVALUE = argDict['GETVALUE'] # gets a specific value from given date
ALTTODAY = argDict['ALTTODAY'] # stores the input date yyyy-mm-dd 
if argDict['GENALL']: # generate all files 
    CSV = True
    FILLPLATFORM = True
    TP = True 

# setting date used throughout execution 
if (ALTTODAY == ''):
    today = date.today()
else:
    # set today date to command line argument value 
    todayList = ALTTODAY[0].split('-')
    today = date(int(todayList[0]), int(todayList[1]), int(todayList[2]))

listDate = str(today).split('-')
TODAYDATE = f'{listDate[1]}/{listDate[2]}'  # mm/yy

STRTODAY = today.strftime('%Y-%m-%d') # used with polygon data; yy-mm-dd
tomorrow = today + timedelta(1)
STRTOMORROW = tomorrow.strftime('%Y-%m-%d') # used for yahoo finance history 
yesterday = today - timedelta(1)
STRYESTERDAY = yesterday.strftime('%Y-%m-%d') 

# Email variables 
EMAILADDRESS = 'etfsender@gmail.com'
EMAILPASSWORD = 'egztwpmmkbicpjfd' # 'P@55w0rd123' 

EMAILLIST = [ 'trallen@davidson.edu', 'michaelgkelly01@yahoo.com', 'ludurkin@davidson.edu', 'hannachrisj@gmail.com' ] 
# EMAILLIST = [ 'trallen@davidson.edu' , 'michaelgkelly01@yahoo.com'] # can be used for testing
#TestPlatformEMAILLIST =  [ 'hannachrisj@gmail.com' ]

# determines if application is a script file or frozen exe
# not sure exactly what this means, found it on stack overflow 
if getattr(sys, 'frozen', False):
    curDir = os.path.dirname(sys.executable)
elif __file__:
    curDir = os.path.abspath(__file__)

# #store the directory part of the aboslute path of the current file
script_dir = os.path.dirname(os.path.abspath(__file__))

# Navigate up the directory tree until src not located to find the root directory
while "src" in script_dir:
    script_dir = os.path.dirname(script_dir)
    
TPROOT = script_dir  #root directory for trading post execution 

# Debug files 
LOGROOT = f'{TPROOT}/debug'
LOGFILE = f'{LOGROOT}/logfile.txt'
try:
    os.mkdir(f'{LOGROOT}')
    logmsg('DEBUG', '008', f'created src directory \'{LOGROOT}\'')
except FileExistsError:
    # trims the log file to preserve space 
    with open(LOGFILE, mode='r+') as lf:
        lastLines = lf.readlines()[-2000:] # how many lines to save from previous log file 
    with open(LOGFILE, mode='w') as lf:
        for line in lastLines:
            lf.write(line)
    logmsg('DEBUG', '009', f'debug directory already created at \'{LOGROOT}\'')

# Template files 
SRCROOT = f'{TPROOT}/src'
TEMPLATEPLATFORM = f'{SRCROOT}/TA.WORK.xlsx' 
TEMPEXCEL = f'{SRCROOT}/stocktradingpostdemo.xlsx' 
try:
    os.mkdir(f'{SRCROOT}')
    logmsg('DEBUG', '004', f'created src directory \'{SRCROOT}\'')
except FileExistsError:
    logmsg('DEBUG', '005', f'src directory already created at \'{SRCROOT}\'')

# Output files 
OUTROOT = f'{TPROOT}'
OUTPUTPLATFORM = f'{OUTROOT}/testTP/{STRTODAY}_testPlatform.xlsx'
OUTPUTEXCEL = f'{OUTROOT}/testTP/{STRTODAY}_testTradingPost.xlsx'
CSVFILE = f'{OUTROOT}/outFiles/{STRTODAY}_csv.csv' 
try:
    os.mkdir(f'{OUTROOT}')
    logmsg('DEBUG', '006', f'created src directory \'{OUTROOT}\'')
except FileExistsError:
    logmsg('DEBUG', '007', f'output directory already created at \'{OUTROOT}\'')

# polygon login 
key = "CP1nN_q8W8C4eG7phIPNgLNCyPEyDZPe" # paid standard version 
try:
    CLIENT = RESTClient(key)
    logmsg('DEBUG', '001', 'loading RESTClient')
except Exception as e:
    logmsg('ERROR', '002', f'{e}')

TICKERS = [ 'JNK', 'GDX', 'VCR', 'VDC', 'VIG', 'VDE', 'VFH', 
        'VWO', 'VHT', 'VIS', 'VGT', 'VAW', 'VNQ', 'VOO', 
        'VOX', 'BND', 'BNDX', 'VXUS', 'VTI', 'VPU', 'XTN' ]

# TICKERS = [ 'JNK' ] # used for testing 

PARAMSET = [[ 'minute', 1 ], # one minute time interval 
            [ 'minute', 5 ], # 5 minute time interval 
            [ 'day', 1 ]] #one day time interval 

INDICATORS = [ 'one_min_50', 'one_min_200', 'five_min_50', 'five_min_200', 'one_day_50', 'one_day_200', 'close_price' ]
MINDICATORS = [ 'five_min_50', 'five_min_200', 'one_min_50', 'one_min_200' ]
DAYDICATORS = [ 'one_day_50', 'one_day_200' ]

# this is { platform column: indicator }
INPUTS = { 'G':'close_price', 'H':'one_day_50', 'I':'one_day_200', 'J':'five_min_50', 
            'K':'five_min_200', 'L':'one_min_50', 'M':'one_min_200' }

PLATFORMCOLS = { 'date':'E', 'close_price':'G', 'one_day_50':'H', 'one_day_200':'I', 'five_min_50':'J', 
            'five_min_200':'K', 'one_min_50':'L', 'one_min_200':'M' }

# base cell for each ticker in Trading Post excel 
ETFBASECELL = { 'JNK':'B3', 'GDX':'C3', 'VCR':'D3', 'VDC':'E3', 'VIG':'F3', 
            'VDE':'G3', 'VFH':'H3', 'VWO':'B13', 'VHT':'C13', 'VIS':'D13', 'VGT':'E13', 
            'VAW':'F13', 'VNQ':'G13', 'VOO':'H13', 'VOX':'B23', 'BND':'C23', 
            'BNDX':'D23', 'VXUS':'E23', 'VTI':'F23', 'VPU':'G23', 'XTN':'H23' }

# cell with date in Trading Post 
PLATDATECELL = 'B3'

try: 
    # loading excel as workbook object
    workbook = openpyxl.load_workbook(TEMPEXCEL)
    excelSheet = workbook.active
except Exception as e:
    print(f'ERROR: {e}')
    sys.exit(4)

color_black = '00000000' # color black 
color_white = 'FFFFFFFF' # color white 
PLAINCOLOR = openpyxl.styles.PatternFill(start_color=color_white, end_color=color_white, fill_type='solid')

# Cell color templates 
dark_green = '064A23' # Buy Signal HexColor
light_green = '87C94B' # Hold-Buy Signal HexColor
brown_yellow = 'A28818' # Sell Signal HexColor
bright_yellow ='F6DD58' # Hold-Sell Signal HexColor
try: 
    BUYCOLOR = openpyxl.styles.PatternFill(start_color = dark_green, end_color = dark_green, fill_type = 'solid')
    logmsg('DEBUG', '010', 'BUY Signal Color Template')
except Exception as e:
    BUYCOLOR = PLAINCOLOR
    logmsg('ERROR', '011', f'{e}')
    logmsg('NOTICE', '011', 'BUY Color Signal set as plain')
try:
    HOBUYCOLOR = openpyxl.styles.PatternFill(start_color = light_green, end_color = light_green, fill_type = 'solid')
    logmsg('DEBUG', '012', 'HOBUY Signal Color Template')
except Exception as e:
    HOBUYCOLOR = PLAINCOLOR
    logmsg('ERROR', '013', f'{e}')
    logmsg('NOTICE', '013', 'HOBUY Color Signal set as plain')
try:
    SELLCOLOR = openpyxl.styles.PatternFill(start_color = brown_yellow, end_color = brown_yellow, fill_type = 'solid')
    logmsg('DEBUG', '014', 'SELL Signal Color Template')
except Exception as e:
    SELLCOLOR = PLAINCOLOR
    logmsg('ERROR', '015', f'{e}')
    logmsg('NOTICE', '015', 'SELL Color Signal set as plain')
try:
    HOSELLCOLOR = openpyxl.styles.PatternFill(start_color = bright_yellow, end_color = bright_yellow, fill_type = 'solid')
    logmsg('DEBUG', '016', 'HOSELL Signal Color Template')
except Exception as e:
    HOSELLCOLOR = PLAINCOLOR
    logmsg('ERROR', '017', f'{e}')
    logmsg('NOTICE', '017', 'HOSELL Color Signal set as plain')