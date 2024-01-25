'''
This file is used to generate the Trading Post, csv or platform files 

log numbers 300-399 
'''

import csv 
import openpyxl
import shutil 
import configuration_file as config

# generate_tp: generates the Trading Post output excel 
# parameters: 
#       etfDict - dictionary, a dictionary of etf objects 
#       outputExcel - string, the path to the output Trading Post excel file 
# returns:  0 - successfully generated Trading Post 
#           1 - failed 
def generate_tp(etfDict, outputExcel):
    # makes a copy of the template Trading Post file
    try:
        shutil.copyfile(config.TEMPEXCEL, outputExcel)
    except FileNotFoundError as e:
        config.logmsg('ERROR', 107, f'{e}')
        # not sure what I meant right here, don't know what this error would mean either 
        config.logmsg('NOTICE', 108, f'creating file {outputExcel}')
        print(f'creating file {outputExcel}') # TODO: Actually create the missing path; I don't remember what this means 
        return 1

    # loading excel as workbook object
    workbook = openpyxl.load_workbook(outputExcel)
    activeSheet = workbook.active

    for ticker in config.TICKERS:
        curEtf = etfDict[ticker]
        curBase = curEtf.basecell
        charBase = curBase[0]
        numBase = int(curBase[1:])

        activeSheet[curBase] = curEtf.ticker # setting ticker name in tp
        # activeSheet[f'{charBase}{numBase + 1}'].value = curEtf.name # setting etf name in tp #### Leave this commented for now 
        activeSheet[f'{charBase}{numBase + 3}'] = config.TODAYDATE # setting today date in tp 

        activeSheet[f'{charBase}{numBase + 5}'] = curEtf.signal # Buy/Sell/Hold signal 
        cF = activeSheet[f'{charBase}{numBase + 5}'].font
        if curEtf.signal == '!BUY!':
            activeSheet[f'{charBase}{numBase + 5}'].font = openpyxl.styles.Font(name=f'{cF.name}', color=f'{config.color_white}', sz=f'{cF.sz}', b=True)
        elif curEtf.signal == '!SELL!':
            activeSheet[f'{charBase}{numBase + 5}'].font = openpyxl.styles.Font(name=f'{cF.name}', color=f'{config.color_black}', sz=f'{cF.sz}', b=True)
        else: 
            activeSheet[f'{charBase}{numBase + 5}'].font = openpyxl.styles.Font(name=f'{cF.name}', color=f'{config.color_black}', sz=f'{cF.sz}')
        activeSheet[f'{charBase}{numBase + 5}'].fill = curEtf.color # Buy/Sell/Hold color 
        activeSheet[f'{charBase}{numBase + 6}'] = curEtf.minTradeRange # min for trade range 
        activeSheet[f'{charBase}{numBase + 7}'] = curEtf.maxTradeRange # max for trade range 
        activeSheet[f'{charBase}{numBase + 8}'] = curEtf.indicatorDict['close_price'] # setting close price in tp 

    config.logmsg('DEBUG', 100, f'saving Trading Post as {outputExcel}')
    workbook.save(outputExcel)
    workbook.close()
    
    return 0

# generate_csv: generates a csv file with path config.CSVFILE
# parameters: 
#       etfDict - dictionary, a dictionary of etf objects 
#       csvFile - string, the path to the output csv file 
# returns:  0 - successfully generated csv file
#           1 - failed 
def generate_csv(etfDict, csvFile):
    try: 
        with open(csvFile, mode='w') as curCsv:
            fieldnames = ['ticker', 'one_day_50',
                            'one_day_200', 'five_min_50', 
                            'five_min_200','one_min_50', 
                            'one_min_200', 'close_price', 'signal']
            writer = csv.DictWriter(curCsv, fieldnames=fieldnames)
            writer.writeheader() # creates the csv header from fieldnames 
            for ticker in config.TICKERS:
                curEtf = etfDict[ticker]
                etfVals = curEtf.indicatorDict
                config.logmsg('DEBUG', 430 + config.TICKERS.index(ticker), f'writing csv row for {ticker}')
                writer.writerow({
                    'ticker': curEtf.ticker,
                    'one_day_50': etfVals['one_day_50'],
                    'one_day_200': etfVals['one_day_200'],
                    'five_min_50': etfVals['five_min_50'],
                    'five_min_200': etfVals['five_min_200'],
                    'one_min_50': etfVals['one_min_50'],
                    'one_min_200': etfVals['one_min_200'],
                    'close_price': etfVals['close_price'], 
                    'signal': curEtf.signal
                })
        config.logmsg('DEBUG', 400, f'saving csv file as {csvFile}')
    except Exception as e:
        config.logmsg('ERROR', 401, f'{e}')
        return 1
    return 0

# fill_platform: fills the excel platform with values  
# parameters: 
#       etfDict - dictionary, a dictionary of etf objects 
#       outputPlatform - string, the path to the output platform excel file 
# returns:  0 - successfully generated the output platform 
#           1 - failed 
def fill_platform(etfDict, outputPlatform):
    tickerRowDict = {} # { 'JNK': 6 }
    
    # makes a copy of the template platform file
    try:
        shutil.copyfile(config.TEMPLATEPLATFORM, outputPlatform)
        config.logmsg('DEBUG', 300, f'copied template platform to {outputPlatform}')
    except Exception as e:
        config.logmsg('ERROR', 301, f'{e}')
        return 1

    # loading excel as workbook object
    try:
        workbook = openpyxl.load_workbook(outputPlatform)
        config.logmsg('DEBUG', 302, f'loaded output platform as excel')
    except Exception as e:
        config.logmsg('ERROR', 303, f'{e}')
        return 1

    activeSheet = workbook.active

    # setting the platform date cell 
    activeSheet[config.PLATDATECELL] = f'{config.TODAYDATE}/{config.listDate[0]}'

    # going through each cell and getting ticker index 
    config.logmsg('DEBUG', 305, 'getting platform rows for each ticker')
    for row in activeSheet.iter_rows(max_row=activeSheet.max_row, max_col=1):
        for cell in row:
            # gets the rows with tickers in the excel 
            if cell.value in config.TICKERS:
                tickerRowDict[cell.value] = cell.coordinate[1:]

    # fills the index values in platform row for each ticker 
    for ticker in config.TICKERS:
        curRow = tickerRowDict[ticker]
        config.logmsg('DEBUG', 330 + config.TICKERS.index(ticker), f'filling platform, row: {curRow} for ticker: {ticker}')
        for col in config.PLATFORMCOLS:
            # TODO: maybe could remove 'date' from PLATFORMCOLS 
            if col == 'date':
                activeSheet[f'{config.PLATFORMCOLS[col]}{curRow}'] = config.TODAYDATE
            else:
                activeSheet[f'{config.PLATFORMCOLS[col]}{curRow}'] = etfDict[ticker].indicatorDict[col]

    try:
        workbook.save(outputPlatform)
        workbook.close()
        config.logmsg('DEBUG', 304, 'saving output platform')
    except Exception as e:
        config.logmsg('ERROR', 305, f'{e}')
        return 1

    return 0
