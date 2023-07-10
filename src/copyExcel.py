'''
This script copies the template excel file from configurationFile.py and 
creates the new excel for current day based on generated csv
'''

import shutil
import openpyxl 
import configurationFile as config
import platFormulas as formula


# TODO19: Implement set_ranges 
def set_ranges(rangeType, etf):
    if rangeType == 'buy':
        rangeMin = formula.buy_min(etf)
        rangeMax = formula.buy_max(etf)
    elif rangeType == 'sell':
        rangeMin = formula.sell_min(etf)
        rangeMax = formula.sell_max(etf)
    else:
        return '', '', 1
    if rangeMin > rangeMax:
        return '', '', 1
    return rangeMin, rangeMax, 0

def determine_buy_sell(etf):
    etfVals = etf.indicatorDict
    signal = None
    color = None
    minTradeRange = None
    maxTradeRange = None 

    if (etfVals['close_price'] > etfVals['one_day_50']):
        # Looking for sell signals 
        if (etfVals['close_price'] > etfVals['one_day_200']):
            minPrice = 1000000000
            for col in config.MINDICATORS:
                minPrice = min(minPrice, etfVals[col])
            if (etfVals['close_price'] < minPrice):
                signal = '!SELL!' 
                color = config.SELLCOLOR
            else:
                signal = 'HOLD' 
                color = config.HOSELLCOLOR
            minTradeRange, maxTradeRange, clearTP = set_ranges('sell', etf)
            if (not clearTP):
                return signal, color, minTradeRange, maxTradeRange
    else:
        # Looking for buy signals 
        if (etfVals['close_price'] < etfVals['one_day_50']):
            maxPrice = -1
            for col in config.MINDICATORS:
                maxPrice = max(maxPrice, etfVals[col])
            if (etfVals['close_price'] > maxPrice):
                signal = '!BUY!' 
                color = config.BUYCOLOR 
            else:
                signal = 'HOLD' 
                color = config.HOBUYCOLOR 
            minTradeRange, maxTradeRange, clearTP = set_ranges('buy', etf)
            if (not clearTP):
                return signal, color, minTradeRange, maxTradeRange
    signal = 'HOLD' 
    color = config.PLAINCOLOR 
    minTradeRange, maxTradeRange, clearTP = set_ranges('', etf)
    return signal, color, minTradeRange, maxTradeRange

def fill_platform(etfDict):

    tickerRowDict = {} # { 'JNK': 6 }

    # makes a copy of the template platform file
    shutil.copyfile(config.TEMPLATEPLATFORM, config.OUTPUTPLATFORM)

    # loading excel as workbook object
    workbook = openpyxl.load_workbook(config.OUTPUTPLATFORM)
    activeSheet = workbook.active

    # going through each cell and getting ticker index 
    for row in activeSheet.iter_rows(max_row=activeSheet.max_row, max_col=1):
        for cell in row:
            # gets the rows with tickers in the excel 
            if cell.value in config.TICKERS:
                tickerRowDict[cell.value] = cell.coordinate[1:]

    for ticker in config.TICKERS:
        for col in config.PLATFORMCOLS:
            if col == 'date':
                activeSheet[f'{config.PLATFORMCOLS[col]}{tickerRowDict[ticker]}'] = config.TODAYDATE
            else:
                activeSheet[f'{config.PLATFORMCOLS[col]}{tickerRowDict[ticker]}'] = etfDict[ticker].indicatorDict[col]

    workbook.save(config.OUTPUTPLATFORM)
    if (config.DEBUG):
        print(f'created temp platform {config.OUTPUTPLATFORM}')

    workbook.close()
